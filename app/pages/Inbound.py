import io
import sys
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import plotly.graph_objects as go

# ============================================================
# 1) CONFIG / CONSTANTS
# ============================================================
st.set_page_config(page_title="Inbound Validator", layout="wide")

# ---- WINDOWS BEEP (REAL) ----
IS_WINDOWS = sys.platform.startswith("win")
if IS_WINDOWS:
    import winsound


def beep_ok():
    if IS_WINDOWS:
        winsound.Beep(1200, 120)


def beep_error():
    if IS_WINDOWS:
        winsound.Beep(300, 180)


def beep_dup():
    if IS_WINDOWS:
        winsound.Beep(700, 90)
        winsound.Beep(700, 90)


# ---- COLUMNAS (internas normalizadas) ----
COL_PEDIDO = "PEDIDO"
COL_MATERIAL = "MATERIAL"
COL_PERFIL = "PERFIL"
COL_SERIAL = "SERIE"
COL_CANTIDAD = "CANTIDAD"

# creadas por la app
COL_SERIE_FIS = "SERIE_FISICA"
COL_QTY = "QTY_FISICA"
COL_VALID = "VALIDACION"

# ---- PERFIL ----
PERFIL_CON_SERIE = "CON PERFIL DE SERIE"
PERFIL_SIN_SERIE = "SIN PERFIL DE SERIE"

# ---- VALIDACIONES OK ----
VALID_SERIAL_OK = "Serial Correcto"
VALID_QTY_OK = "Cantidad Correcta"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


# ============================================================
# 2) UTILS: NORMALIZE / CLEAN
# ============================================================
def normalize_header(s: str) -> str:
    """Normaliza headers para que coincidan aunque vengan en mayúsculas/minúsculas/tildes/espacios."""
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    for ch in [" ", "-", ".", "/", "\\"]:
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s


def canonicalize_df_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Renombra headers a forma normalizada y mapea el serial esperado a COL_SERIAL=SERIE."""
    df = df.rename(columns={c: normalize_header(c) for c in df.columns})

    serial_candidates = [
        "NUMERO_DE_SERIE_FABRICANTE",
        "NUMERO_SERIE_FABRICANTE",
        "NUMERO_DE_SERIE",
        "SERIAL",
        "SERIE",
    ]
    found = None
    for cand in serial_candidates:
        if cand in df.columns:
            found = cand
            break
    if found and found != COL_SERIAL:
        df = df.rename(columns={found: COL_SERIAL})

    return df


def norm(s) -> str:
    if pd.isna(s):
        return ""
    return str(s).strip()


def ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Valida columnas mínimas del archivo y crea las columnas que maneja la app."""
    required = [COL_PEDIDO, COL_MATERIAL, COL_PERFIL, COL_CANTIDAD]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"The File is Missing Several Required Columns: {missing}")

    # Crear si no existen
    if COL_SERIAL not in df.columns:
        df[COL_SERIAL] = ""
    if COL_SERIE_FIS not in df.columns:
        df[COL_SERIE_FIS] = ""
    if COL_QTY not in df.columns:
        df[COL_QTY] = pd.NA
    if COL_VALID not in df.columns:
        df[COL_VALID] = ""

    # Tipos / limpieza
    df[COL_PEDIDO] = df[COL_PEDIDO].astype(str).str.strip()
    df[COL_MATERIAL] = df[COL_MATERIAL].astype(str).str.strip()
    df[COL_PERFIL] = df[COL_PERFIL].astype(str).str.strip()

    df[COL_SERIAL] = df[COL_SERIAL].astype(str).replace({"nan": ""}).str.strip()
    df[COL_SERIE_FIS] = df[COL_SERIE_FIS].astype(str).replace({"nan": ""}).str.strip()

    df[COL_CANTIDAD] = pd.to_numeric(df[COL_CANTIDAD], errors="coerce").fillna(0)
    df[COL_QTY] = pd.to_numeric(df[COL_QTY], errors="coerce")
    df[COL_VALID] = df[COL_VALID].astype(str).replace({"nan": ""}).str.strip()

    return df


def is_ok(val: str, ok_text: str) -> bool:
    return norm(val).lower() == ok_text.lower()


# ============================================================
# 3) BUSINESS LOGIC: SERIAL / QTY
# ============================================================
def build_serial_index(df_scope: pd.DataFrame):
    """
    serial_map: serial esperado -> [indices]
    pending: indices pendientes (no validados como Serial Correcto)
    """
    serial_map = {}
    pending = set()
    for idx, row in df_scope.iterrows():
        expected = norm(row.get(COL_SERIAL, ""))
        if not expected:
            continue
        serial_map.setdefault(expected, []).append(idx)
        if not is_ok(row.get(COL_VALID, ""), VALID_SERIAL_OK):
            pending.add(idx)
    return serial_map, pending


def expected_qty_scope(df_scope: pd.DataFrame) -> float:
    return float(df_scope[COL_CANTIDAD].sum())


# ============================================================
# 4) EXCEL EXPORT
# ============================================================
def save_excel_with_styles(original_bytes: bytes, df_updated: pd.DataFrame) -> bytes:
    """
    Escribe QTY_FISICA, VALIDACION, SERIE_FISICA al archivo original.
    Si columnas no existen, las crea como nuevas columnas.
    """
    wb = load_workbook(io.BytesIO(original_bytes))
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[normalize_header(v)] = col

    def ensure_excel_col(norm_name: str, display_name: str) -> int:
        if norm_name in headers:
            return headers[norm_name]
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col).value = display_name
        headers[norm_name] = new_col
        return new_col

    col_qty = ensure_excel_col(COL_QTY, "QTY FISICA")
    col_val = ensure_excel_col(COL_VALID, "VALIDACIÓN")
    col_sf = ensure_excel_col(COL_SERIE_FIS, "SERIE FISICA")

    for df_idx, row in df_updated.iterrows():
        excel_row = int(df_idx) + 2
        ws.cell(row=excel_row, column=col_qty).value = None if pd.isna(row.get(COL_QTY)) else row.get(COL_QTY)
        ws.cell(row=excel_row, column=col_val).value = row.get(COL_VALID, "")
        ws.cell(row=excel_row, column=col_sf).value = row.get(COL_SERIE_FIS, "")

        v = norm(row.get(COL_VALID, "")).lower()
        if v in (VALID_SERIAL_OK.lower(), VALID_QTY_OK.lower()):
            ws.cell(row=excel_row, column=col_val).fill = GREEN_FILL

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ============================================================
# 5) SUMMARY + GAUGES (SCOPE)
# ============================================================
def scope_key(order_id: str, material: str) -> str:
    return f"{order_id}||{material}"


def compute_summary(df_scope_now: pd.DataFrame, perfil_sel: str, skey: str):
    """
    Retorna:
    Pending Serialization, Pending Quantity,
    Verified Serialization, Verified Quantity,
    Invalid Serialization, Invalid Quantity
    """
    pending_serial = pending_qty = 0
    verified_serial = verified_qty = 0
    invalid_serial = invalid_qty = 0

    if perfil_sel.strip().upper() == PERFIL_CON_SERIE.upper():
        _, pending_set = build_serial_index(df_scope_now)
        pending_serial = len(pending_set)

        verified_serial = int(
            df_scope_now[COL_VALID].astype(str).str.strip().str.lower().eq(VALID_SERIAL_OK.lower()).sum()
        )

        invalid_nf = st.session_state.get("invalid_serial_by_scope", {}).get(skey, 0)
        invalid_dup = st.session_state.get("dup_serial_by_scope", {}).get(skey, 0)
        invalid_serial = int(invalid_nf + invalid_dup)

    elif perfil_sel.strip().upper() == PERFIL_SIN_SERIE.upper():
        qty_has_value = df_scope_now[COL_QTY].notna().any()
        pending_qty = 0 if qty_has_value else 1

        verified_qty = 1 if df_scope_now[COL_VALID].astype(str).str.strip().str.lower().eq(VALID_QTY_OK.lower()).any() else 0

        if qty_has_value and verified_qty == 0:
            invalid_qty = 1

    return pending_serial, pending_qty, verified_serial, verified_qty, invalid_serial, invalid_qty


def progress_serial_pct(df_scope_now: pd.DataFrame) -> float:
    total_expected = int(df_scope_now[COL_SERIAL].astype(str).str.strip().ne("").sum())
    verified = int(df_scope_now[COL_VALID].astype(str).str.strip().str.lower().eq(VALID_SERIAL_OK.lower()).sum())
    return (verified / total_expected * 100.0) if total_expected > 0 else 0.0


def progress_qty_pct(df_scope_now: pd.DataFrame) -> float:
    qty_has_value = df_scope_now[COL_QTY].notna().any()
    if not qty_has_value:
        return 0.0
    ok = df_scope_now[COL_VALID].astype(str).str.strip().str.lower().eq(VALID_QTY_OK.lower()).any()
    return 100.0 if ok else 50.0


def gauge(value_pct: float, title: str):
    """
    Velocímetro limpio. Puedes ajustar rangos/colores después si quieres.
    """
    value_pct = max(0.0, min(100.0, float(value_pct)))
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=value_pct,
        number={"suffix": "%"},
        title={"text": title},
        gauge={
            "axis": {"range": [0, 100]},
            "bar": {"thickness": 0.25},
        }
    ))
    fig.update_layout(height=220, margin=dict(l=10, r=10, t=50, b=0))
    st.plotly_chart(fig, use_container_width=True)


# ============================================================
# 6) APP UI
# ============================================================
st.title("Inbound – Validation (Order → Material → Serie → Quantity )")

uploaded = st.file_uploader("Load Excel", type=["xlsx"])
if not uploaded:
    st.info("Load Excel.")
    st.stop()

file_key = f"{uploaded.name}::{uploaded.size}"
original_bytes = uploaded.getvalue()

# ---- INIT DF (persistente) ----
if st.session_state.get("file_key") != file_key:
    try:
        raw_df = pd.read_excel(io.BytesIO(original_bytes))
        df_init = canonicalize_df_columns(raw_df)
        df_init = ensure_columns(df_init)
    except Exception as e:
        st.error(f"The Excel File Could Not Be Read: {e}")
        st.stop()

    st.session_state["df"] = df_init
    st.session_state["file_key"] = file_key

    # estados
    st.session_state["scan_input"] = ""
    st.session_state["qty_input"] = 0.0
    st.session_state["last_ok"] = ""

    # contadores globales (se mantienen)
    st.session_state["found"] = 0
    st.session_state["not_found"] = 0
    st.session_state["duplicates"] = 0

    # contadores por scope (para Invalid Serialization real por pedido+material)
    st.session_state["invalid_serial_by_scope"] = {}  # not_found por scope
    st.session_state["dup_serial_by_scope"] = {}      # duplicates por scope

df = st.session_state["df"]

# ---- SELECTORES ----
pedidos = sorted(df[COL_PEDIDO].dropna().unique().tolist())
pedido_sel = st.selectbox("1) Select Order", pedidos)

df_pedido = df[df[COL_PEDIDO] == pedido_sel]
materiales = sorted(df_pedido[COL_MATERIAL].dropna().unique().tolist())
material_sel = st.selectbox("2) Select Material", materiales)

mask_scope = (df[COL_PEDIDO] == pedido_sel) & (df[COL_MATERIAL] == material_sel)
df_scope0 = df[mask_scope].copy()

perfiles = df_scope0[COL_PERFIL].dropna().astype(str).str.strip().unique().tolist()
perfil_sel = perfiles[0] if perfiles else ""
if len(perfiles) > 1:
    st.warning(f"⚠️ There is more than one profil with this Order+Material: {perfiles}. It will be used: {perfil_sel}")

skey = scope_key(pedido_sel, material_sel)

# ============================================================
# 7) VALIDATION AREA
# ============================================================
st.subheader("Validation")

if st.session_state.get("last_ok"):
    st.info(f"✅ Last OK: **{st.session_state['last_ok']}**")

st.caption(f"ORDER: {pedido_sel}  |  MATERIAL: {material_sel}  |  PROFIL: {perfil_sel}")

# ---- PERFIL CON SERIE ----
if perfil_sel.strip().upper() == PERFIL_CON_SERIE.upper():
    df_scope = df[mask_scope].copy()

    if df_scope[COL_SERIAL].astype(str).str.strip().eq("").all():
        st.error(f"PROFIL='{PERFIL_CON_SERIE}' There are not series in the column.")
        st.stop()

    serial_map, pending = build_serial_index(df_scope)
    disabled_scan = (len(pending) == 0)
    if disabled_scan:
        st.success("✅ There are not pending serial numbers for this Order + Material.")

    def process_scan():
        serial_scanned = st.session_state.scan_input.strip()
        if not serial_scanned:
            return

        df_scope_local = df[mask_scope].copy()
        serial_map_local, pending_local = build_serial_index(df_scope_local)

        if serial_scanned in serial_map_local:
            row_to_mark = None
            for idx in serial_map_local[serial_scanned]:
                if idx in pending_local:
                    row_to_mark = idx
                    break

            if row_to_mark is None:
                # DUPLICADO
                st.session_state["duplicates"] += 1
                st.session_state["dup_serial_by_scope"][skey] = st.session_state["dup_serial_by_scope"].get(skey, 0) + 1
                beep_dup()
            else:
                # OK
                df.at[row_to_mark, COL_QTY] = 1
                df.at[row_to_mark, COL_VALID] = VALID_SERIAL_OK
                df.at[row_to_mark, COL_SERIE_FIS] = serial_scanned

                st.session_state["found"] += 1
                st.session_state["last_ok"] = serial_scanned
                beep_ok()
        else:
            # INVALIDO (NO EXISTE)
            st.session_state["not_found"] += 1
            st.session_state["invalid_serial_by_scope"][skey] = st.session_state["invalid_serial_by_scope"].get(skey, 0) + 1
            beep_error()

        st.session_state.scan_input = ""
        st.session_state["df"] = df
        st.rerun()

    st.text_input(
        "Scan the serial number:",
        key="scan_input",
        placeholder="Scan one serial number…",
        disabled=disabled_scan,
        on_change=process_scan
    )

# ---- PERFIL SIN SERIE ----
elif perfil_sel.strip().upper() == PERFIL_SIN_SERIE.upper():
    df_scope = df[mask_scope].copy()
    expected = expected_qty_scope(df_scope)
    st.write(f"📦 **Expected Quantity (QUANTITY): {expected:g}**")

    st.number_input(
        "Input Quantity Received:",
        min_value=0.0,
        step=1.0,
        key="qty_input"
    )

    def apply_qty():
        entered = float(st.session_state.get("qty_input", 0) or 0)
        diff = entered - expected

        if abs(diff) < 1e-9:
            msg = VALID_QTY_OK
            beep_ok()
            st.session_state["last_ok"] = f"{material_sel} → {entered:g}"
        elif diff < 0:
            msg = f"Lower Quantity than Expected {abs(diff):g}"
            beep_error()
        else:
            msg = f"Higher Quantity than Expected {diff:g}"
            beep_error()

        df.loc[mask_scope, COL_QTY] = entered
        df.loc[mask_scope, COL_VALID] = msg

        st.session_state["df"] = df
        st.rerun()

    st.button("Apply Quantity", on_click=apply_qty)

else:
    st.error(f"PERFIL no reconocido: '{perfil_sel}'.")
    st.stop()

# ============================================================
# 8) SUMMARY (HORIZONTAL) + GAUGES
# ============================================================
st.divider()

df_scope_now = st.session_state["df"][mask_scope].copy()
pending_ser, pending_qty, verified_ser, verified_qty, invalid_ser, invalid_qty = compute_summary(df_scope_now, perfil_sel, skey)

# ---- SUMMARY ROW ----
s1, s2, s3, s4, s5, s6 = st.columns([1.2, 1.2, 1.2, 1.2, 1.35, 1.2])
with s1:
    st.metric("Pending Serialization", pending_ser)
with s2:
    st.metric("Pending Quantity", pending_qty)
with s3:
    st.metric("Verified Serialization", verified_ser)
with s4:
    st.metric("Verified Quantity", verified_qty)
with s5:
    st.metric("Invalid Serialization", invalid_ser)
with s6:
    st.metric("Invalid Quantity", invalid_qty)

# ---- GAUGES ROW (below summary) ----
g1, g2 = st.columns(2)

with g1:
    if perfil_sel.strip().upper() == PERFIL_CON_SERIE.upper():
        pct = progress_serial_pct(df_scope_now)
        gauge(pct, "Serialization Progress")
        st.progress(min(max(pct / 100.0, 0), 1))
    else:
        gauge(0, "Serialization Progress (N/A)")
        st.progress(0.0)

with g2:
    if perfil_sel.strip().upper() == PERFIL_SIN_SERIE.upper():
        pct = progress_qty_pct(df_scope_now)
        gauge(pct, "Quantity Progress")
        st.progress(min(max(pct / 100.0, 0), 1))
    else:
        gauge(0, "Quantity Progress (N/A)")
        st.progress(0.0)

# ============================================================
# 9) TABLE
# ============================================================
st.divider()
st.subheader("TABLE")

df_scope_view = st.session_state["df"][mask_scope].copy()
view_cols = [COL_PEDIDO, COL_MATERIAL, COL_PERFIL, COL_SERIAL, COL_CANTIDAD, COL_SERIE_FIS, COL_QTY, COL_VALID]
st.dataframe(df_scope_view[view_cols], use_container_width=True)

# ============================================================
# 10) DOWNLOADS
# ============================================================
st.divider()

if st.button("Generate Updated Excel"):
    try:
        out_bytes = save_excel_with_styles(original_bytes, st.session_state["df"])
        st.download_button(
            "Download Updated Excel",
            data=out_bytes,
            file_name="Formato_Verificacion_ACTUALIZADO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success("Excel generated with QTY FISICA, VALIDACIÓN y SERIE FISICA.")
    except Exception as e:
        st.error(f"Error generating Excel: {e}")

report_cols = [COL_PEDIDO, COL_MATERIAL, COL_PERFIL, COL_SERIAL, COL_CANTIDAD, COL_SERIE_FIS, COL_QTY, COL_VALID]
df_report = st.session_state["df"][report_cols].copy()

rep_bytes = io.BytesIO()
with pd.ExcelWriter(rep_bytes, engine="openpyxl") as writer:
    df_report.to_excel(writer, index=False, sheet_name="REPORTE")
rep_bytes.seek(0)

st.download_button(
    "Download REPORT (QTY_FISICA + VALIDACION + SERIE_FISICA)",
    data=rep_bytes.getvalue(),
    file_name="REPORTE_VERIFICACION.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

